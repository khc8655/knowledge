import re
from typing import List, Dict, Any, Optional
from .base import PipelineBase, PipelineError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class ExcelPipeline(PipelineBase):
    SOURCE_TYPE = "excel"
    SOURCE_ID = "04"

    MODEL_PATTERN = re.compile(r'[A-Z]{2,}\d+(?:\s*[A-Z]+)?(?:\s*V\d+\.?\d*)?')

    def parse(self, file_path: str) -> List[Dict[str, Any]]:
        profile = self.profile(file_path)
        config = self.default_config(profile)
        return self.generate_cards(file_path, config)

    def profile(self, file_path: str) -> List[Dict]:
        if load_workbook is None:
            raise PipelineError("openpyxl not installed")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise PipelineError(f"Excel 解析失败: {e}")

        results = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(max_row=20, values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])

            if not rows:
                continue

            header_row = self._detect_header_row(rows)
            cols = rows[header_row] if header_row < len(rows) else []
            num_cols = len(cols)

            model_cols = []
            price_cols = []
            note_cols = []

            for ci, col_name in enumerate(cols):
                col_values = [rows[r][ci] if ci < len(rows[r]) else "" for r in range(header_row + 1, min(len(rows), header_row + 20))]

                if self.MODEL_PATTERN.search(col_name) or sum(1 for v in col_values if self.MODEL_PATTERN.search(v)) >= max(1, len(col_values) * 0.3):
                    model_cols.append(ci)

                if any(kw in col_name for kw in ["价", "金额", "price", "cost", "Price", "Cost"]):
                    price_cols.append(ci)

                if ci >= num_cols - 3:
                    avg_len = sum(len(str(v)) for v in col_values) / max(1, len(col_values))
                    if avg_len > 10:
                        note_cols.append(ci)

            discontinued_cols = []
            for ci in range(num_cols):
                col_values = [rows[r][ci] if ci < len(rows[r]) else "" for r in range(header_row + 1, min(len(rows), header_row + 20))]
                if any(kw in str(v) for v in col_values for kw in ["停产", "替代", "停售", "EOL"]):
                    discontinued_cols.append(ci)

            results.append({
                "sheet_name": ws.title,
                "header_row": header_row,
                "skip_rows": header_row,
                "model_cols": model_cols,
                "price_cols": price_cols,
                "title_cols": model_cols,
                "body_cols": [i for i in range(num_cols) if i not in model_cols],
                "keyword_cols": [],
                "note_cols": note_cols,
                "discontinued_cols": discontinued_cols,
                "comparison_mode": False,
                "num_rows": len(rows) - header_row - 1,
                "num_cols": num_cols,
            })

        wb.close()
        return results

    def default_config(self, profile: List[Dict]) -> List[Dict]:
        return [{
            "sheet_name": s["sheet_name"],
            "header_row": s["header_row"],
            "skip_rows": s["skip_rows"],
            "title_cols": s["title_cols"],
            "body_cols": s["body_cols"],
            "model_cols": s["model_cols"],
            "price_cols": s["price_cols"],
            "note_cols": s["note_cols"],
            "discontinued_cols": s["discontinued_cols"],
            "comparison_mode": s["comparison_mode"],
        } for s in profile]

    def generate_cards(self, file_path: str, config: List[Dict]) -> List[Dict]:
        if load_workbook is None:
            raise PipelineError("openpyxl not installed")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise PipelineError(f"Excel 解析失败: {e}")

        doc_file = file_path.split("/")[-1].split("\\")[-1]
        slug = self.doc_slug(doc_file)
        cards = []
        seq = 0

        for sheet_cfg in config:
            sheet_name = sheet_cfg["sheet_name"]
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            header_row = sheet_cfg["header_row"]
            title_cols = sheet_cfg["title_cols"]
            body_cols = sheet_cfg["body_cols"]

            rows = list(ws.iter_rows(values_only=True))

            for ri in range(header_row + 1, len(rows)):
                row = rows[ri]
                row_values = [str(c) if c is not None else "" for c in row]

                if all(not v.strip() for v in row_values):
                    continue

                title_parts = [row_values[ci].strip() for ci in title_cols if ci < len(row_values) and row_values[ci].strip()]
                title = " | ".join(title_parts) if title_parts else f"行{ri}"

                body_parts = [row_values[ci].strip() for ci in body_cols if ci < len(row_values) and row_values[ci].strip()]
                body = " | ".join(body_parts) if body_parts else ""

                if not body:
                    continue

                path = f"{doc_file} > {sheet_name}"
                cards.append(self.make_card(
                    doc_file=doc_file, doc_slug=slug, title=title,
                    level=0, path=path, line_start=ri, body=body, seq=seq,
                ))
                seq += 1

        wb.close()
        return cards

    def _detect_header_row(self, rows: List[List[str]]) -> int:
        best_row = 0
        best_score = 0
        for ri in range(min(10, len(rows))):
            row = rows[ri]
            score = sum(2 for cell in row if str(cell).strip() and len(str(cell).strip()) < 20 and not any(c.isdigit() for c in str(cell)))
            score += sum(1 for cell in row if str(cell).strip() and len(str(cell).strip()) < 30)
            if score > best_score:
                best_score = score
                best_row = ri
        return best_row
